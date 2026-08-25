"""Audit a local KIMORE dataset and build a subject/exercise manifest.

The script never modifies the dataset. It reads folder names, clinical-score
workbooks, and skeletal CSV files, then writes:

* ``kimore_manifest.csv`` -- one row per subject/exercise
* ``kimore_audit_summary.txt`` -- dataset counts and detected problems
* ``first_valid_sample.png`` -- a quick sanity plot for one recording

Usage (PowerShell):

    python kimore_dataset_audit.py "C:\\KIMORE"

Dependencies:

    python -m pip install numpy matplotlib openpyxl
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
import warnings
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import numpy as np
from openpyxl import load_workbook


warnings.filterwarnings(
    "ignore",
    message="Unknown extension is not supported and will be removed",
    module=r"openpyxl\..*",
)


SUBJECT_PATTERN = re.compile(r"^(E|NE|B|P|S)_ID(\d+)$", re.IGNORECASE)
EXERCISE_PATTERN = re.compile(r"^Es([1-5])$", re.IGNORECASE)
SCORE_HEADER_PATTERN = re.compile(
    r"clinical\s*(TS|PO|CF)\s*Ex\s*#?\s*([1-5])", re.IGNORECASE
)

COHORTS = {
    "E": "expert_control",
    "NE": "nonexpert_control",
    "B": "back_pain",
    "P": "parkinson",
    "S": "stroke",
}

KINECT_JOINTS = (
    "SpineBase",
    "SpineMid",
    "Neck",
    "Head",
    "ShoulderLeft",
    "ElbowLeft",
    "WristLeft",
    "HandLeft",
    "ShoulderRight",
    "ElbowRight",
    "WristRight",
    "HandRight",
    "HipLeft",
    "KneeLeft",
    "AnkleLeft",
    "FootLeft",
    "HipRight",
    "KneeRight",
    "AnkleRight",
    "FootRight",
    "SpineShoulder",
    "HandTipLeft",
    "ThumbLeft",
    "HandTipRight",
    "ThumbRight",
)

MANIFEST_FIELDS = (
    "sample_id",
    "subject_id",
    "cohort",
    "exercise",
    "clinical_ts",
    "clinical_po",
    "clinical_cf",
    "position_frames",
    "orientation_frames",
    "timestamp_frames",
    "position_columns",
    "orientation_columns",
    "position_path",
    "orientation_path",
    "timestamp_path",
    "position_target_usable",
    "status",
    "issues",
    "warnings",
)


@dataclass(frozen=True)
class CsvAudit:
    rows: int
    columns: int | None
    inconsistent_rows: int
    nonnumeric_rows: int


def numeric_suffix(path: Path) -> tuple[str, int]:
    match = SUBJECT_PATTERN.fullmatch(path.name)
    if not match:
        return path.name, sys.maxsize
    return match.group(1).upper(), int(match.group(2))


def find_subject_directories(root: Path) -> list[Path]:
    """Find subject folders without depending on a specific top-level layout."""
    found: dict[str, Path] = {}
    for candidate in root.rglob("*"):
        match = SUBJECT_PATTERN.fullmatch(candidate.name)
        if not candidate.is_dir() or not match:
            continue
        if not any(EXERCISE_PATTERN.fullmatch(child.name) for child in candidate.iterdir() if child.is_dir()):
            continue
        subject_id = candidate.name.upper()
        if subject_id in found and found[subject_id] != candidate:
            raise RuntimeError(
                f"Duplicate subject folder for {subject_id}: {found[subject_id]} and {candidate}"
            )
        found[subject_id] = candidate
    return sorted(found.values(), key=numeric_suffix)


def trimmed_row(row: list[str]) -> list[str]:
    while row and not row[-1].strip():
        row.pop()
    return row


def audit_numeric_csv(path: Path) -> CsvAudit:
    rows = 0
    expected_columns: int | None = None
    inconsistent_rows = 0
    nonnumeric_rows = 0

    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        for raw_row in csv.reader(handle):
            row = trimmed_row(raw_row)
            if not row:
                continue
            rows += 1
            if expected_columns is None:
                expected_columns = len(row)
            elif len(row) != expected_columns:
                inconsistent_rows += 1
            try:
                for value in row:
                    float(value)
            except ValueError:
                nonnumeric_rows += 1

    return CsvAudit(rows, expected_columns, inconsistent_rows, nonnumeric_rows)


def read_numeric_csv(path: Path) -> np.ndarray:
    rows: list[list[float]] = []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        for raw_row in csv.reader(handle):
            row = trimmed_row(raw_row)
            if row:
                rows.append([float(value) for value in row])
    if not rows:
        raise ValueError(f"No numeric rows found in {path}")
    return np.asarray(rows, dtype=np.float64)


def find_exactly_one(folder: Path, pattern: str, issues: list[str]) -> Path | None:
    matches = sorted(folder.glob(pattern)) if folder.exists() else []
    if len(matches) == 1:
        return matches[0]
    if not matches:
        issues.append(f"missing {pattern}")
    else:
        issues.append(f"multiple {pattern} files")
    return None


def _read_scores_from_workbook(
    workbook_path: Path,
) -> tuple[dict[tuple[str, int], float], dict[int, list[str]]]:
    """Read all score fields while keeping problems local to their exercise."""
    issues_by_exercise: dict[int, list[str]] = {
        exercise: [] for exercise in range(1, 6)
    }
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    sheet = workbook.active
    values = list(sheet.iter_rows(values_only=True))
    workbook.close()
    if len(values) < 2:
        message = f"clinical workbook has fewer than two rows: {workbook_path}"
        return {}, {exercise: [message] for exercise in range(1, 6)}

    headers = values[0]
    data = values[1]
    scores: dict[tuple[str, int], float] = {}
    for index, header in enumerate(headers):
        if header is None:
            continue
        match = SCORE_HEADER_PATTERN.search(str(header))
        if not match:
            continue
        kind = match.group(1).lower()
        exercise = int(match.group(2))
        value = data[index] if index < len(data) else None
        if value is None:
            issues_by_exercise[exercise].append(
                f"missing {kind.upper()} score for Es{exercise}"
            )
            continue
        try:
            scores[(kind, exercise)] = float(value)
        except (TypeError, ValueError):
            issues_by_exercise[exercise].append(
                f"nonnumeric {kind.upper()} score for Es{exercise}: {value!r}"
            )

    for exercise in range(1, 6):
        if ("ts", exercise) not in scores:
            issues_by_exercise[exercise].append(f"TS score not found for Es{exercise}")
    return scores, issues_by_exercise


def read_subject_scores(
    subject_dir: Path,
) -> tuple[dict[tuple[str, int], float], dict[int, list[str]], list[str]]:
    """Read consistent, correctly named workbook copies for one subject.

    KIMORE normally repeats the same subject-level workbook below every
    exercise.  A misfiled workbook from another subject must never become the
    target source merely because its path sorts first.
    """
    workbooks = sorted(subject_dir.glob("Es*/Label/ClinicalAssessment_*.xlsx"))
    if not workbooks:
        issue = "missing clinical assessment workbook"
        return {}, {exercise: [issue] for exercise in range(1, 6)}, []

    expected_filename = f"ClinicalAssessment_{subject_dir.name}.xlsx"
    expected_pattern = re.compile(
        rf"^ClinicalAssessment_{re.escape(subject_dir.name)}(?:\(\d+\))?\.xlsx$",
        re.IGNORECASE,
    )
    expected_workbooks = [
        path for path in workbooks if expected_pattern.fullmatch(path.name)
    ]
    unexpected_workbooks = [
        path for path in workbooks if not expected_pattern.fullmatch(path.name)
    ]
    warnings_found: list[str] = []
    if unexpected_workbooks:
        relative_paths = ", ".join(
            str(path.relative_to(subject_dir)) for path in unexpected_workbooks
        )
        warnings_found.append(
            "unexpected clinical workbook filename(s), ignored: " + relative_paths
        )
    if not expected_workbooks:
        issue = f"missing correctly named {expected_filename} workbook"
        return {}, {exercise: [issue] for exercise in range(1, 6)}, warnings_found

    parsed = [_read_scores_from_workbook(path) for path in expected_workbooks]
    canonical_scores, canonical_issues = parsed[0]
    if any(scores != canonical_scores for scores, _ in parsed[1:]):
        relative_paths = ", ".join(
            str(path.relative_to(subject_dir)) for path in expected_workbooks
        )
        issue = "conflicting correctly named clinical workbook copies: " + relative_paths
        return {}, {exercise: [issue] for exercise in range(1, 6)}, warnings_found

    issues_by_exercise = {
        exercise: list(canonical_issues[exercise]) for exercise in range(1, 6)
    }
    for _, workbook_issues in parsed[1:]:
        for exercise in range(1, 6):
            issues_by_exercise[exercise].extend(workbook_issues[exercise])
            issues_by_exercise[exercise] = list(
                dict.fromkeys(issues_by_exercise[exercise])
            )
    return canonical_scores, issues_by_exercise, warnings_found


def audit_exercise(
    subject_dir: Path,
    subject_id: str,
    cohort: str,
    exercise_number: int,
    scores: dict[tuple[str, int], float],
    exercise_score_issues: Iterable[str],
    subject_warnings: Iterable[str],
) -> dict[str, object]:
    exercise_name = f"Es{exercise_number}"
    exercise_dir = subject_dir / exercise_name
    raw_dir = exercise_dir / "Raw"
    issues = list(exercise_score_issues)

    if not exercise_dir.exists():
        issues.append(f"missing {exercise_name} folder")

    position_path = find_exactly_one(raw_dir, "JointPosition*.csv", issues)
    orientation_path = find_exactly_one(raw_dir, "JointOrientation*.csv", issues)
    timestamp_path = find_exactly_one(raw_dir, "TimeStamp*.csv", issues)

    audits: dict[str, CsvAudit | None] = {
        "position": audit_numeric_csv(position_path) if position_path else None,
        "orientation": audit_numeric_csv(orientation_path) if orientation_path else None,
        "timestamp": audit_numeric_csv(timestamp_path) if timestamp_path else None,
    }

    position = audits["position"]
    orientation = audits["orientation"]
    timestamp = audits["timestamp"]

    for name, audit in audits.items():
        if audit is None:
            continue
        if audit.rows == 0:
            issues.append(f"empty {name} CSV")
        if audit.inconsistent_rows:
            issues.append(f"{name} has {audit.inconsistent_rows} inconsistent-width rows")
        if audit.nonnumeric_rows:
            issues.append(f"{name} has {audit.nonnumeric_rows} nonnumeric rows")

    if position and position.columns != 100:
        issues.append(f"position has {position.columns} columns after trimming, expected 100")
    if orientation and orientation.columns != 100:
        issues.append(f"orientation has {orientation.columns} columns after trimming, expected 100")
    if timestamp and timestamp.columns != 1:
        issues.append(f"timestamp has {timestamp.columns} columns after trimming, expected 1")

    frame_counts = [
        audit.rows for audit in (position, orientation, timestamp) if audit is not None
    ]
    if len(frame_counts) > 1 and len(set(frame_counts)) != 1:
        issues.append(f"frame-count mismatch: {frame_counts}")

    if ("ts", exercise_number) not in scores:
        issues.append("missing target TS score")

    position_target_usable = bool(
        ("ts", exercise_number) in scores
        and position is not None
        and position.rows > 0
        and position.columns == 100
        and position.inconsistent_rows == 0
        and position.nonnumeric_rows == 0
    )

    return {
        "sample_id": f"{subject_id}_{exercise_name}",
        "subject_id": subject_id,
        "cohort": cohort,
        "exercise": exercise_name,
        "clinical_ts": scores.get(("ts", exercise_number), ""),
        "clinical_po": scores.get(("po", exercise_number), ""),
        "clinical_cf": scores.get(("cf", exercise_number), ""),
        "position_frames": position.rows if position else "",
        "orientation_frames": orientation.rows if orientation else "",
        "timestamp_frames": timestamp.rows if timestamp else "",
        "position_columns": position.columns if position else "",
        "orientation_columns": orientation.columns if orientation else "",
        "position_path": str(position_path.resolve()) if position_path else "",
        "orientation_path": str(orientation_path.resolve()) if orientation_path else "",
        "timestamp_path": str(timestamp_path.resolve()) if timestamp_path else "",
        "position_target_usable": str(position_target_usable).lower(),
        "status": "ok" if not issues else "problem",
        "issues": "; ".join(dict.fromkeys(issues)),
        "warnings": "; ".join(dict.fromkeys(subject_warnings)),
    }


def build_manifest(root: Path) -> list[dict[str, object]]:
    subject_dirs = find_subject_directories(root)
    if not subject_dirs:
        raise RuntimeError(
            f"No KIMORE subject folders found beneath {root}. "
            "Expected names such as E_ID1, NE_ID1, B_ID1, P_ID1, or S_ID1."
        )

    rows: list[dict[str, object]] = []
    for subject_dir in subject_dirs:
        match = SUBJECT_PATTERN.fullmatch(subject_dir.name)
        assert match is not None
        prefix = match.group(1).upper()
        subject_id = subject_dir.name.upper()
        scores, score_issues_by_exercise, subject_warnings = read_subject_scores(
            subject_dir
        )
        for exercise_number in range(1, 6):
            rows.append(
                audit_exercise(
                    subject_dir,
                    subject_id,
                    COHORTS[prefix],
                    exercise_number,
                    scores,
                    score_issues_by_exercise[exercise_number],
                    subject_warnings,
                )
            )
    return rows


def write_manifest(rows: list[dict[str, object]], path: Path) -> None:
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=MANIFEST_FIELDS)
        writer.writeheader()
        writer.writerows(rows)


def write_summary(rows: list[dict[str, object]], path: Path) -> None:
    subjects = {str(row["subject_id"]) for row in rows}
    cohort_subjects: dict[str, set[str]] = {}
    for row in rows:
        cohort_subjects.setdefault(str(row["cohort"]), set()).add(str(row["subject_id"]))

    status_counts = Counter(str(row["status"]) for row in rows)
    exercise_counts = Counter(
        str(row["exercise"]) for row in rows if row["status"] == "ok"
    )
    usable_position_counts = Counter(
        str(row["exercise"])
        for row in rows
        if str(row["position_target_usable"]).casefold() == "true"
    )
    warnings_by_subject = {
        str(row["subject_id"]): str(row["warnings"])
        for row in rows
        if row.get("warnings")
    }
    scores_by_exercise: dict[str, list[float]] = {}
    for row in rows:
        value = row["clinical_ts"]
        if value != "":
            scores_by_exercise.setdefault(str(row["exercise"]), []).append(float(value))

    lines = [
        "KIMORE DATASET AUDIT",
        "====================",
        f"Subjects found: {len(subjects)}",
        f"Subject/exercise rows: {len(rows)}",
        f"Rows fully OK: {status_counts['ok']}",
        f"Rows with audit problems: {status_counts['problem']}",
        f"Subjects with non-blocking warnings: {len(warnings_by_subject)}",
        "",
        "Subjects by cohort:",
    ]
    for cohort in sorted(cohort_subjects):
        lines.append(f"  {cohort}: {len(cohort_subjects[cohort])}")

    lines.extend(["", "Fully OK recordings by exercise:"])
    for exercise in ("Es1", "Es2", "Es3", "Es4", "Es5"):
        lines.append(f"  {exercise}: {exercise_counts[exercise]}")

    lines.extend(["", "Position+target usable recordings by exercise:"])
    for exercise in ("Es1", "Es2", "Es3", "Es4", "Es5"):
        lines.append(f"  {exercise}: {usable_position_counts[exercise]}")

    lines.extend(["", "Clinical TS score distribution by exercise:"])
    for exercise in ("Es1", "Es2", "Es3", "Es4", "Es5"):
        values = scores_by_exercise.get(exercise, [])
        if not values:
            lines.append(f"  {exercise}: no scores")
            continue
        array = np.asarray(values, dtype=float)
        lines.append(
            f"  {exercise}: n={len(array)}, min={array.min():.3f}, "
            f"median={np.median(array):.3f}, max={array.max():.3f}, "
            f"mean={array.mean():.3f}, std={array.std(ddof=0):.3f}"
        )

    problem_rows = [row for row in rows if row["status"] != "ok"]
    lines.extend(["", "Problems:"])
    if not problem_rows:
        lines.append("  None detected.")
    else:
        for row in problem_rows:
            lines.append(f"  {row['sample_id']}: {row['issues']}")

    lines.extend(["", "Warnings:"])
    if not warnings_by_subject:
        lines.append("  None detected.")
    else:
        for subject_id, warning in sorted(warnings_by_subject.items()):
            lines.append(f"  {subject_id}: {warning}")

    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def save_first_sample_plot(rows: list[dict[str, object]], path: Path) -> str | None:
    first = next(
        (
            row
            for row in rows
            if row["status"] == "ok" and row["position_path"] and row["clinical_ts"] != ""
        ),
        None,
    )
    if first is None:
        return None

    import matplotlib.pyplot as plt

    raw = read_numeric_csv(Path(str(first["position_path"])))
    positions = raw.reshape(raw.shape[0], len(KINECT_JOINTS), 4)
    spine_base_xyz = positions[:, 0, :3]

    figure, axis = plt.subplots(figsize=(10, 4.8))
    for coordinate, label in enumerate(("X", "Y", "Z")):
        axis.plot(spine_base_xyz[:, coordinate], label=label, linewidth=1)
    axis.set_title(
        f"{first['sample_id']} – SpineBase position – clinical TS {float(first['clinical_ts']):.3f}"
    )
    axis.set_xlabel("Frame")
    axis.set_ylabel("Kinect coordinate")
    axis.legend()
    axis.grid(alpha=0.25)
    figure.tight_layout()
    figure.savefig(path, dpi=160)
    plt.close(figure)
    return (
        f"First valid sample: {first['sample_id']} | "
        f"position shape={positions.shape} | clinical TS={float(first['clinical_ts']):.3f}"
    )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("kimore_root", type=Path, help="Path to the extracted KIMORE folder")
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path("kimore_audit_output"),
        help="Directory for generated outputs (default: ./kimore_audit_output)",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    root = args.kimore_root.expanduser().resolve()
    output_dir = args.output_dir.expanduser().resolve()

    if not root.is_dir():
        print(f"Error: KIMORE root is not a directory: {root}", file=sys.stderr)
        return 2

    output_dir.mkdir(parents=True, exist_ok=True)
    print(f"Scanning: {root}")
    rows = build_manifest(root)

    manifest_path = output_dir / "kimore_manifest.csv"
    summary_path = output_dir / "kimore_audit_summary.txt"
    plot_path = output_dir / "first_valid_sample.png"
    write_manifest(rows, manifest_path)
    write_summary(rows, summary_path)
    first_sample_message = save_first_sample_plot(rows, plot_path)

    print(summary_path.read_text(encoding="utf-8"))
    if first_sample_message:
        print(first_sample_message)
    print(f"Manifest: {manifest_path}")
    print(f"Summary:  {summary_path}")
    if plot_path.exists():
        print(f"Plot:     {plot_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
